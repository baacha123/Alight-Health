#!/usr/bin/env python3
"""
GovCloud Agent Evaluation (with Auto-Patching)
===============================================
All-in-one script for running IBM watsonx Orchestrate evaluations on GovCloud/FedRAMP
environments where certain models (like 405b) are not available.

This script automatically:
1. Patches the agentops package to work with GovCloud models
2. Generates the ADK-compatible eval_config.yml
3. Runs the evaluation
4. Provides LLM-as-Judge evaluation

Usage:
    python govcloud_eval.py --setup             # Verify/apply patches only
    python govcloud_eval.py --generate          # Generate test cases from Excel
    python govcloud_eval.py --evaluate          # Run ADK evaluation
    python govcloud_eval.py --judge             # Run LLM-as-Judge on results
    python govcloud_eval.py --all               # Generate + Evaluate + Judge + Report
    python govcloud_eval.py --report            # Show results from last run

Requirements:
    - ADK venv activated with `orchestrate` CLI available
    - WXO environment authenticated
    - govcloud_config.yaml configured with available models
"""

import json
import os
import sys
import re
import csv
import shutil
import argparse
import subprocess
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, Optional, List
import importlib.util

# =============================================================================
# CONFIGURATION LOADING
# =============================================================================

SCRIPT_DIR = Path(__file__).parent
DEFAULT_CONFIG_FILE = SCRIPT_DIR / "govcloud_config.yaml"


def load_yaml_config(config_path: Path) -> Dict[str, Any]:
    """Load YAML config file."""
    try:
        import yaml
    except ImportError:
        print("ERROR: PyYAML required. Run: pip install pyyaml")
        sys.exit(1)

    if not config_path.exists():
        print(f"ERROR: Config file not found: {config_path}")
        print("Create govcloud_config.yaml or specify with --config")
        sys.exit(1)

    with open(config_path) as f:
        return yaml.safe_load(f)


def load_env_file():
    """Load environment variables from .env file."""
    env_paths = [
        Path(__file__).parent.parent / "config" / ".env",
        Path(__file__).parent / "eval.env",
        Path(__file__).parent / ".env",
    ]
    for env_path in env_paths:
        if env_path.exists():
            with open(env_path) as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith("#") and "=" in line:
                        key, _, value = line.partition("=")
                        key = key.strip()
                        value = value.strip().strip('"').strip("'")
                        if key and value and key not in os.environ:
                            os.environ[key] = value
            break


# Load .env early
load_env_file()


# =============================================================================
# AGENTOPS PATCHING SYSTEM
# =============================================================================

def find_agentops_package() -> Optional[Path]:
    """Find the agentops package in the current environment."""
    try:
        spec = importlib.util.find_spec("agentops")
        if spec and spec.origin:
            return Path(spec.origin).parent
    except (ImportError, AttributeError):
        pass

    # Fallback: search in site-packages
    for path in sys.path:
        agentops_path = Path(path) / "agentops"
        if agentops_path.exists() and (agentops_path / "__init__.py").exists():
            return agentops_path

    return None


def get_patch_definitions(config: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    """
    Define all patches needed for GovCloud compatibility.
    Returns dict of {filename: {patches}}
    """
    llm_judge_model = config.get("models", {}).get("llm_judge", "meta-llama/llama-3-2-90b-vision-instruct")

    return {
        "main.py": {
            "description": "Disable hardcoded Langfuse telemetry",
            "patches": [
                {
                    "find": 'os.environ["TELEMETRY_PLATFORM"] = "langfuse"',
                    "replace": '# os.environ["TELEMETRY_PLATFORM"] = "langfuse"  # GOVCLOUD: Disabled',
                    "required": False,  # May already be patched
                },
                {
                    "find": "exporters=[LangfusePersistence()],",
                    "replace": "exporters=[],  # GOVCLOUD: Disabled Langfuse",
                    "required": False,
                },
            ],
        },
        "clients.py": {
            "description": "Handle dict config for custom_metrics_config",
            "patches": [
                {
                    "find": "llamaj_config_dict[\"model_id\"] = cmc.llmaaj_config.model_id",
                    "replace": """# ============ GOVCLOUD FIX: Handle both dict and dataclass ============
    cmc = config.custom_metrics_config
    if isinstance(cmc, dict):
        llmaaj_cfg = cmc.get("llmaaj_config", {})
        llamaj_config_dict["model_id"] = llmaaj_cfg.get("model_id", llamaj_config_dict["model_id"])
        llamaj_config_dict["embedding_model_id"] = llmaaj_cfg.get("embedding_model_id", llamaj_config_dict["embedding_model_id"])
    else:
        llamaj_config_dict["model_id"] = cmc.llmaaj_config.model_id
        llamaj_config_dict["embedding_model_id"] = cmc.llmaaj_config.embedding_model_id
    # ============ END GOVCLOUD FIX ============""",
                    "required": False,
                    "check_already_patched": "GOVCLOUD FIX: Handle both dict and dataclass",
                },
            ],
        },
        "runner.py": {
            "description": "Handle dict config and pass llmaaj_provider",
            "patches": [
                {
                    "find": "if config.extractors_config.paths is not None:",
                    "replace": """# ============ GOVCLOUD FIX: Handle both dict and dataclass ============
    # Load custom extractors
    ext_cfg = config.extractors_config
    ext_paths = ext_cfg.get("paths") if isinstance(ext_cfg, dict) else ext_cfg.paths
    if ext_paths is not None:""",
                    "required": False,
                    "check_already_patched": "GOVCLOUD FIX: Handle both dict and dataclass",
                },
                {
                    "find": "if config.custom_metrics_config.paths is not None:",
                    "replace": """# Load custom evaluations
    cmc = config.custom_metrics_config
    cmc_paths = cmc.get("paths") if isinstance(cmc, dict) else cmc.paths
    if cmc_paths is not None:""",
                    "required": False,
                    "check_already_patched": "cmc_paths = cmc.get",
                },
                {
                    "find": "custom_llmaaj_client=None,",
                    "replace": "custom_llmaaj_client=llmaaj_provider,  # GOVCLOUD: Pass provider",
                    "required": False,
                    "check_already_patched": "custom_llmaaj_client=llmaaj_provider",
                },
            ],
        },
        "evaluation_package.py": {
            "description": "Use custom LLM client instead of hardcoded 405b",
            "patches": [
                {
                    # This is a more complex patch - we need to add the conditional
                    "find": 'llm_client=get_provider(\n                model_id="meta-llama/llama-3-405b-instruct",',
                    "replace": f"""# ============ GOVCLOUD FIX: Use custom_llmaaj_client if provided ============
        if custom_llmaaj_client:
            llm_client_for_eval = custom_llmaaj_client
        else:
            llm_client_for_eval = get_provider(
                model_id="{llm_judge_model}",""",
                    "required": False,
                    "check_already_patched": "GOVCLOUD FIX: Use custom_llmaaj_client",
                },
            ],
        },
    }


def check_file_patched(file_path: Path, patch_info: Dict) -> bool:
    """Check if a file already has the patch applied."""
    if not file_path.exists():
        return False

    content = file_path.read_text(encoding='utf-8')

    for patch in patch_info.get("patches", []):
        check_str = patch.get("check_already_patched", patch["replace"][:50])
        if check_str in content:
            return True

    return False


def apply_patch_to_file(file_path: Path, patch_info: Dict, dry_run: bool = False) -> Dict[str, Any]:
    """Apply patches to a single file."""
    result = {
        "file": str(file_path),
        "status": "unknown",
        "message": "",
    }

    if not file_path.exists():
        result["status"] = "error"
        result["message"] = "File not found"
        return result

    # Check if already patched
    if check_file_patched(file_path, patch_info):
        result["status"] = "already_patched"
        result["message"] = "Already patched"
        return result

    content = file_path.read_text(encoding='utf-8')
    original_content = content
    patches_applied = 0

    for patch in patch_info.get("patches", []):
        find_str = patch["find"]
        replace_str = patch["replace"]

        if find_str in content:
            content = content.replace(find_str, replace_str, 1)
            patches_applied += 1

    if patches_applied == 0:
        result["status"] = "no_match"
        result["message"] = "Pattern not found (may be different version)"
        return result

    if dry_run:
        result["status"] = "would_patch"
        result["message"] = f"Would apply {patches_applied} patch(es)"
        return result

    # Backup and write
    backup_path = file_path.with_suffix(file_path.suffix + ".govcloud.bak")
    if not backup_path.exists():
        shutil.copy2(file_path, backup_path)

    file_path.write_text(content, encoding='utf-8')

    # Clear pycache
    pycache_dir = file_path.parent / "__pycache__"
    if pycache_dir.exists():
        shutil.rmtree(pycache_dir, ignore_errors=True)

    result["status"] = "patched"
    result["message"] = f"Applied {patches_applied} patch(es)"
    return result


def apply_all_patches(config: Dict[str, Any], dry_run: bool = False) -> bool:
    """Apply all patches to the agentops package."""
    print(f"\n{'='*60}")
    print("  GOVCLOUD PATCH SYSTEM")
    print(f"{'='*60}")

    agentops_path = find_agentops_package()
    if not agentops_path:
        print("\nERROR: Could not find agentops package.")
        print("Make sure ibm-watsonx-orchestrate-evaluation-framework is installed.")
        return False

    print(f"\nFound agentops at: {agentops_path}")
    if dry_run:
        print("(DRY RUN - no changes will be made)")

    patch_defs = get_patch_definitions(config)
    all_success = True
    results = []

    for filename, patch_info in patch_defs.items():
        file_path = agentops_path / filename
        result = apply_patch_to_file(file_path, patch_info, dry_run)
        results.append(result)

        status_symbol = {
            "patched": "\u2705",
            "already_patched": "\u2714\ufe0f",
            "would_patch": "\U0001f504",
            "no_match": "\u26a0\ufe0f",
            "error": "\u274c",
        }.get(result["status"], "?")

        print(f"\n  {status_symbol} {filename}: {result['message']}")
        print(f"     {patch_info['description']}")

        if result["status"] == "error":
            all_success = False

    print(f"\n{'─'*60}")

    patched_count = sum(1 for r in results if r["status"] in ["patched", "already_patched"])
    print(f"  Patches verified: {patched_count}/{len(patch_defs)}")

    if all_success:
        print("  Status: READY for GovCloud evaluation")
    else:
        print("  Status: Some patches could not be applied")

    print(f"{'─'*60}\n")

    return all_success


def verify_patches(config: Dict[str, Any]) -> bool:
    """Verify all patches are applied."""
    agentops_path = find_agentops_package()
    if not agentops_path:
        return False

    patch_defs = get_patch_definitions(config)
    all_patched = True

    for filename, patch_info in patch_defs.items():
        file_path = agentops_path / filename
        if not check_file_patched(file_path, patch_info):
            all_patched = False

    return all_patched


# =============================================================================
# EVAL CONFIG GENERATION
# =============================================================================

def generate_eval_config(config: Dict[str, Any], output_path: Path) -> Path:
    """Generate the ADK-compatible eval_config.yml from govcloud_config.yaml."""
    try:
        import yaml
    except ImportError:
        print("ERROR: PyYAML required. Run: pip install pyyaml")
        sys.exit(1)

    models = config.get("models", {})
    paths = config.get("paths", {})
    auth = config.get("auth", {})
    provider_cfg = config.get("provider", {})
    eval_cfg = config.get("evaluation", {})

    # Build ADK config structure
    adk_config = {
        "auth_config": {
            "url": auth.get("url") or os.getenv("WXO_INSTANCE_URL", ""),
            "tenant_name": auth.get("tenant_name", ""),
            "token": auth.get("api_key") or os.getenv("WXO_API_KEY", ""),
        },
        "test_paths": [paths.get("test_cases", "./eval_data_ground_truth")],
        "output_dir": paths.get("output_dir", "./eval_results_govcloud"),
        "llm_user_config": {
            "model_id": models.get("llm_user", "meta-llama/llama-3-2-90b-vision-instruct"),
            "prompt_config": {},
            "user_response_style": "",
        },
        "provider_config": {
            "model_id": models.get("llm_user", "meta-llama/llama-3-2-90b-vision-instruct"),
            "provider": provider_cfg.get("type", "gateway"),
            "embedding_model_id": models.get("embedding", "sentence-transformers/all-minilm-l6-v2"),
            "vendor": provider_cfg.get("vendor", "ibm"),
            "referenceless_eval": False,
        },
        "custom_metrics_config": {
            "llmaaj_config": {
                "model_id": models.get("llm_judge", "meta-llama/llama-3-2-90b-vision-instruct"),
                "provider": provider_cfg.get("type", "gateway"),
                "embedding_model_id": models.get("embedding", "sentence-transformers/all-minilm-l6-v2"),
                "vendor": provider_cfg.get("vendor", "ibm"),
                "referenceless_eval": False,
            },
            "paths": None,
        },
        "extractors_config": {
            "paths": None,
        },
        "metrics": ["JourneySuccessMetric", "ToolCalling"],
        "num_workers": eval_cfg.get("num_workers", 1),
        "n_runs": eval_cfg.get("n_runs", 1),
        "similarity_threshold": eval_cfg.get("similarity_threshold", 0.8),
        "enable_fuzzy_matching": eval_cfg.get("enable_fuzzy_matching", False),
        "is_strict": eval_cfg.get("is_strict", True),
        "skip_legacy_evaluation": False,  # Use legacy to avoid Langfuse
        "data_annotation_run": False,
        "enable_recursive_search": True,
        "tags": None,
        "error_keywords": None,
    }

    with open(output_path, "w") as f:
        yaml.safe_dump(adk_config, f, default_flow_style=False, sort_keys=False)

    return output_path


# =============================================================================
# OPENPYXL / TEST GENERATION (from run_eval_at_scale.py)
# =============================================================================

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# IBM watsonx.ai SDK
try:
    from ibm_watsonx_ai import Credentials
    from ibm_watsonx_ai.foundation_models import ModelInference
    WATSONX_AVAILABLE = True
except ImportError:
    WATSONX_AVAILABLE = False


# Keyword extraction rules
KEYWORD_RULES = {
    "deductible": ["deductible"],
    "copay": ["copay"],
    "coinsurance": ["coinsurance"],
    "premium": ["premium"],
    "out-of-pocket": ["out-of-pocket", "maximum"],
    "in-network": ["in-network", "network"],
    "out-of-network": ["out-of-network"],
    "preventive": ["preventive"],
    "emergency": ["emergency"],
    "specialist": ["specialist"],
    "prescription": ["prescription"],
    "mental health": ["mental health"],
    "dental": ["dental"],
    "vision": ["vision"],
    "disability": ["disability"],
    "fmla": ["FMLA", "leave"],
    "cobra": ["COBRA", "continuation"],
    "hsa": ["HSA", "savings"],
    "fsa": ["FSA", "flexible"],
    "enrollment": ["enrollment", "enroll"],
    "coverage": ["coverage", "covered"],
    "benefit": ["benefit"],
    "plan": ["plan"],
}


def extract_keywords(question: str) -> list:
    """Extract expected response keywords from the question text."""
    question_lower = question.lower()
    keywords = []

    for term, kws in KEYWORD_RULES.items():
        if term in question_lower:
            keywords.extend(kws)

    seen = set()
    unique = []
    for kw in keywords:
        if kw.lower() not in seen:
            seen.add(kw.lower())
            unique.append(kw)

    if not unique:
        stop_words = {"what", "how", "is", "are", "do", "does", "can", "will",
                      "the", "my", "i", "a", "an", "to", "for", "of", "in"}
        words = question_lower.replace("?", "").replace(",", "").split()
        unique = [w for w in words if w not in stop_words and len(w) > 3][:3]

    return unique[:4]


def generate_test_case(question: str, config: Dict[str, Any], expected_response: str = "") -> dict:
    """Generate a ground truth JSON test case from a question."""
    gen_cfg = config.get("generation", {})
    agent_name = gen_cfg.get("agent_name", "alight_supervisor_agent")
    tool_name = gen_cfg.get("tool_name", "call_verint_studio_for_hr_and_benefits_questions")

    keywords = extract_keywords(question)

    return {
        "agent": agent_name,
        "goals": {
            f"{tool_name}-1": ["summarize"]
        },
        "goal_details": [
            {
                "type": "tool_call",
                "name": f"{tool_name}-1",
                "tool_name": tool_name,
                "args": {
                    "query": question
                }
            },
            {
                "name": "summarize",
                "type": "text",
                "response": expected_response,
                "keywords": keywords
            }
        ],
        "story": f"You want to know about your health benefits. Specifically: {question}",
        "starting_sentence": question
    }


def load_questions_from_excel(excel_path: Path, mvp_only: bool = True) -> list:
    """Load health questions from the Excel file."""
    if not OPENPYXL_AVAILABLE:
        print("ERROR: openpyxl required. Run: pip install openpyxl")
        sys.exit(1)

    if not excel_path.exists():
        print(f"ERROR: Excel file not found: {excel_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path, read_only=True)
    questions = []

    if "Health Questions" in wb.sheetnames:
        ws = wb["Health Questions"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if row and len(row) >= 3 and row[2]:
                questions.append(str(row[2]).strip())

    for sheet_name in wb.sheetnames:
        if "nav" in sheet_name.lower():
            ws = wb[sheet_name]
            header = [str(c.value).lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

            q_col = None
            mvp_col = None
            for i, h in enumerate(header):
                if "question" in h and q_col is None:
                    q_col = i
                if "mvp" in h:
                    mvp_col = i

            if q_col is not None:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and len(row) > q_col and row[q_col]:
                        if mvp_only and mvp_col is not None:
                            if row[mvp_col] and str(row[mvp_col]).strip().lower() in ("yes", "y", "true", "1"):
                                questions.append(str(row[q_col]).strip())
                        else:
                            questions.append(str(row[q_col]).strip())

    wb.close()

    seen = set()
    unique_questions = []
    for q in questions:
        if q.lower() not in seen and len(q) > 5:
            seen.add(q.lower())
            unique_questions.append(q)

    return unique_questions


def load_ground_truth(ground_truth_dir: Path) -> Dict[str, Dict]:
    """Load ground truth files that have expected answers."""
    ground_truth = {}

    if not ground_truth_dir.exists():
        return ground_truth

    for gt_file in ground_truth_dir.glob("*.json"):
        try:
            with open(gt_file) as f:
                data = json.load(f)

            question = data.get("starting_sentence", "")
            if not question:
                continue

            expected_response = ""
            keywords = []
            for detail in data.get("goal_details", []):
                if detail.get("type") == "text" or detail.get("name") == "summarize":
                    expected_response = detail.get("response", "")
                    keywords = detail.get("keywords", [])
                    break

            ground_truth[question.lower()] = {
                "question": question,
                "expected_response": expected_response,
                "keywords": keywords,
                "source_file": gt_file.name,
            }
        except Exception as e:
            print(f"  Warning: Could not load {gt_file.name}: {e}")

    return ground_truth


# =============================================================================
# LLM AS JUDGE
# =============================================================================

def get_watsonx_model(config: Dict[str, Any]):
    """Initialize WatsonX model for LLM judge."""
    if not WATSONX_AVAILABLE:
        raise ImportError("ibm-watsonx-ai package not installed. Run: pip install ibm-watsonx-ai")

    api_key = os.getenv("WXO_API_KEY") or os.getenv("WATSONX_API_KEY")
    if not api_key:
        raise ValueError("WXO_API_KEY environment variable not set.")

    credentials = Credentials(
        url="https://us-south.ml.cloud.ibm.com",
        api_key=api_key
    )

    model_id = config.get("models", {}).get("llm_judge", "meta-llama/llama-3-3-70b-instruct")
    project_id = config.get("provider", {}).get("project_id", "skills-flow")

    model = ModelInference(
        model_id=model_id,
        credentials=credentials,
        project_id=project_id
    )
    return model


def llm_judge_evaluate(
    question: str,
    expected_answer: str,
    agent_response: str,
    config: Dict[str, Any]
) -> Dict[str, Any]:
    """Use LLM as a Judge to evaluate if agent response is correct."""
    judge_prompt = f"""You are an expert evaluation judge for a health benefits Q&A system.
Your task is to evaluate if an AI agent's response correctly answers a question about employee benefits.

QUESTION:
{question}

EXPECTED ANSWER (Ground Truth):
{expected_answer}

AGENT'S ACTUAL RESPONSE:
{agent_response}

Respond with ONLY a JSON object:
{{
    "verdict": "CORRECT" or "PARTIALLY_CORRECT" or "INCORRECT",
    "score": <float between 0.0 and 1.0>,
    "reasoning": "<brief explanation>"
}}
"""

    try:
        model = get_watsonx_model(config)
        params = {
            "decoding_method": "greedy",
            "max_new_tokens": 500,
            "temperature": 0,
        }
        result_text = model.generate_text(prompt=judge_prompt, params=params)

        # Parse JSON
        json_start = result_text.find("{")
        json_end = result_text.rfind("}") + 1
        if json_start >= 0 and json_end > json_start:
            result_text = result_text[json_start:json_end]

        result = json.loads(result_text)

        return {
            "success": True,
            "verdict": result.get("verdict", "INCORRECT"),
            "score": float(result.get("score", 0)),
            "reasoning": result.get("reasoning", ""),
            "model": config.get("models", {}).get("llm_judge", "unknown"),
        }

    except Exception as e:
        return {
            "success": False,
            "verdict": "ERROR",
            "score": 0.0,
            "reasoning": f"LLM Judge error: {str(e)}",
        }


# =============================================================================
# COMMAND HANDLERS
# =============================================================================

def cmd_setup(config: Dict[str, Any], args) -> int:
    """Verify and apply patches."""
    success = apply_all_patches(config, dry_run=args.dry_run)
    return 0 if success else 1


def cmd_generate(config: Dict[str, Any], args) -> int:
    """Generate test cases from Excel."""
    print(f"\n{'='*60}")
    print("  GENERATING TEST CASES")
    print(f"{'='*60}")

    paths = config.get("paths", {})
    excel_path = Path(paths.get("excel_file", "./questions.xlsx"))
    output_dir = Path(paths.get("test_cases", "./eval_data_ground_truth"))

    questions = load_questions_from_excel(excel_path, mvp_only=not args.all_questions)
    print(f"\nLoaded {len(questions)} questions from Excel")

    if args.limit:
        questions = questions[:args.limit]
        print(f"Limited to {args.limit} questions")

    output_dir.mkdir(parents=True, exist_ok=True)

    for f in output_dir.glob("health_scale_*.json"):
        f.unlink()

    generated = 0
    for i, question in enumerate(questions, 1):
        test_case = generate_test_case(question, config)
        filename = output_dir / f"health_scale_{i:04d}.json"
        with open(filename, "w") as f:
            json.dump(test_case, f, indent=2)
        generated += 1

    print(f"Generated {generated} test cases in: {output_dir}")
    return 0


def ensure_orchestrate_env(config: Dict[str, Any]) -> bool:
    """Ensure orchestrate environment is activated."""
    auth = config.get("auth", {})
    api_key = auth.get("api_key") or os.getenv("WXO_API_KEY") or os.getenv("WATSONX_API_KEY")

    if not api_key:
        print("\nWARNING: No API key found. Set WXO_API_KEY environment variable.")
        print("Or run: orchestrate env activate <env-name> --api-key <key>")
        return False

    # Check if orchestrate env is active by running a quick command
    result = subprocess.run(
        ["orchestrate", "env", "list"],
        capture_output=True,
        text=True
    )

    if "(active)" not in result.stdout:
        print("\nNo active orchestrate environment. Please run:")
        print("  orchestrate env activate <env-name> --api-key <your-key>")
        return False

    return True


def cmd_evaluate(config: Dict[str, Any], args) -> int:
    """Run ADK evaluation."""
    print(f"\n{'='*60}")
    print("  RUNNING GOVCLOUD EVALUATION")
    print(f"{'='*60}")

    # Check orchestrate environment
    if not ensure_orchestrate_env(config):
        print("\nTIP: Run this first:")
        print("  orchestrate env activate <env-name> --api-key <your-key>")
        return 1

    # Apply patches first
    if not verify_patches(config):
        print("\nPatches not applied. Applying now...")
        if not apply_all_patches(config):
            print("ERROR: Could not apply patches")
            return 1

    # Generate eval config
    eval_config_path = SCRIPT_DIR / "eval_config_generated.yml"
    generate_eval_config(config, eval_config_path)
    print(f"\nGenerated ADK config: {eval_config_path}")

    paths = config.get("paths", {})
    test_dir = Path(paths.get("test_cases", "./eval_data_ground_truth"))
    output_dir = Path(paths.get("output_dir", "./eval_results_govcloud"))

    # Get test files
    test_files = sorted(test_dir.glob("*.json"))
    if args.limit:
        test_files = test_files[:args.limit]

    if not test_files:
        print(f"ERROR: No test files found in {test_dir}")
        return 1

    print(f"Running {len(test_files)} test case(s)")
    output_dir.mkdir(parents=True, exist_ok=True)

    # Build test paths string (comma-separated for multiple files)
    test_paths_str = ",".join(str(f) for f in test_files)

    # Use command-line args instead of --config (more reliable)
    cmd = [
        "orchestrate", "evaluations", "evaluate",
        "--test-paths", test_paths_str,
        "--output-dir", str(output_dir),
    ]

    print(f"\nRunning: {' '.join(cmd)}")
    print(f"{'─'*60}")

    start_time = datetime.now()
    result = subprocess.run(cmd)
    duration = (datetime.now() - start_time).total_seconds()

    print(f"{'─'*60}")
    print(f"Completed in {duration:.1f}s (exit code: {result.returncode})")

    return result.returncode


def cmd_judge(config: Dict[str, Any], args) -> int:
    """Run LLM-as-Judge evaluation."""
    print(f"\n{'='*60}")
    print("  LLM AS JUDGE EVALUATION")
    print(f"{'='*60}")

    paths = config.get("paths", {})
    ground_truth_dir = Path(paths.get("test_cases", "./eval_data_ground_truth"))
    results_dir = Path(paths.get("output_dir", "./eval_results_govcloud"))
    judge_results_file = results_dir / "llm_judge_results.json"

    print(f"\nLLM Judge Model: {config.get('models', {}).get('llm_judge', 'unknown')}")

    ground_truth = load_ground_truth(ground_truth_dir)
    print(f"Loaded {len(ground_truth)} ground truth entries")

    if not ground_truth:
        print("\nWARNING: No ground truth files found.")
        return 1

    # Find messages directory
    messages_dirs = list(results_dir.rglob("messages"))
    if not messages_dirs:
        print("\nERROR: No evaluation results found. Run --evaluate first.")
        return 1

    messages_dir = sorted(messages_dirs)[-1]
    print(f"Reading results from: {messages_dir}")

    judge_results = []
    processed = 0

    for result_file in sorted(messages_dir.glob("*.json")):
        if "metrics" in result_file.name or "analyze" in result_file.name:
            continue

        try:
            with open(result_file) as f:
                result_data = json.load(f)

            messages = result_data if isinstance(result_data, list) else result_data.get("messages", [])

            question = ""
            agent_response = ""

            for msg in messages:
                if not isinstance(msg, dict):
                    continue
                if msg.get("role") == "user" and not question:
                    question = msg.get("content", "")
                elif msg.get("role") == "assistant" and msg.get("type") == "text" and not agent_response:
                    content = msg.get("content", "")
                    if content:
                        agent_response = content

            if not question or not agent_response:
                continue

            gt = ground_truth.get(question.lower())

            if gt and gt.get("expected_response"):
                print(f"\n  Evaluating: {question[:50]}...")

                judge_result = llm_judge_evaluate(
                    question=question,
                    expected_answer=gt["expected_response"],
                    agent_response=agent_response,
                    config=config
                )

                result_entry = {
                    "question": question,
                    "expected_answer": gt["expected_response"][:200],
                    "agent_response": agent_response[:200],
                    "llm_judge": judge_result,
                    "source_file": result_file.name,
                }

                judge_results.append(result_entry)
                processed += 1

                verdict = judge_result.get("verdict", "ERROR")
                score = judge_result.get("score", 0)
                symbol = "OK" if verdict == "CORRECT" else "PARTIAL" if verdict == "PARTIALLY_CORRECT" else "FAIL"
                print(f"    [{symbol}] {verdict} (score: {score:.2f})")

        except Exception as e:
            print(f"  Error processing {result_file.name}: {e}")

        if args.limit and processed >= args.limit:
            break

    # Save results
    judge_results_file.parent.mkdir(parents=True, exist_ok=True)
    with open(judge_results_file, "w") as f:
        json.dump({
            "timestamp": datetime.now().isoformat(),
            "model": config.get("models", {}).get("llm_judge", "unknown"),
            "total_evaluated": len(judge_results),
            "results": judge_results
        }, f, indent=2)

    print(f"\n{'─'*60}")
    print(f"Processed {processed} results")
    print(f"Results saved to: {judge_results_file}")

    if judge_results:
        correct = sum(1 for r in judge_results if r["llm_judge"].get("verdict") == "CORRECT")
        partial = sum(1 for r in judge_results if r["llm_judge"].get("verdict") == "PARTIALLY_CORRECT")
        incorrect = sum(1 for r in judge_results if r["llm_judge"].get("verdict") == "INCORRECT")
        avg_score = sum(r["llm_judge"].get("score", 0) for r in judge_results) / len(judge_results)

        print(f"\n  LLM JUDGE SUMMARY")
        print(f"  {'─'*40}")
        print(f"  Total:     {len(judge_results)}")
        print(f"  Correct:   {correct} ({correct/len(judge_results)*100:.1f}%)")
        print(f"  Partial:   {partial}")
        print(f"  Incorrect: {incorrect}")
        print(f"  Avg Score: {avg_score:.2f}")

    return 0


def cmd_report(config: Dict[str, Any], args) -> int:
    """Show evaluation report."""
    print(f"\n{'='*60}")
    print("  EVALUATION REPORT")
    print(f"{'='*60}")

    paths = config.get("paths", {})
    results_dir = Path(paths.get("output_dir", "./eval_results_govcloud"))

    # Find summary_metrics.csv
    csv_files = list(results_dir.rglob("summary_metrics.csv"))
    if not csv_files:
        print("ERROR: No results found. Run --evaluate first.")
        return 1

    csv_file = sorted(csv_files)[-1]
    print(f"\nResults: {csv_file}")

    with open(csv_file) as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    total = len(rows)
    passed = sum(1 for r in rows if r.get("is_success", "").lower() == "true")
    pass_rate = (passed / total * 100) if total > 0 else 0

    print(f"\n  RESULTS")
    print(f"  {'─'*40}")
    print(f"  Total:  {total}")
    print(f"  Passed: {passed} ({pass_rate:.1f}%)")
    print(f"  Failed: {total - passed}")
    print(f"  {'─'*40}")
    print(f"\n  PASS RATE: {pass_rate:.0f}%")

    # Check for LLM judge results
    judge_file = results_dir / "llm_judge_results.json"
    if judge_file.exists():
        with open(judge_file) as f:
            judge_data = json.load(f)
        results = judge_data.get("results", [])
        if results:
            avg_score = sum(r["llm_judge"].get("score", 0) for r in results) / len(results)
            print(f"\n  LLM JUDGE SCORE: {avg_score*100:.0f}%")

    return 0


# =============================================================================
# MAIN
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="GovCloud Agent Evaluation (with Auto-Patching)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python govcloud_eval.py --setup              # Verify/apply patches
  python govcloud_eval.py --evaluate           # Run evaluation
  python govcloud_eval.py --evaluate --limit 1 # Test with 1 case
  python govcloud_eval.py --judge              # LLM-as-Judge evaluation
  python govcloud_eval.py --all                # Full pipeline
  python govcloud_eval.py --report             # Show results
        """
    )

    parser.add_argument("--config", type=str, default=str(DEFAULT_CONFIG_FILE),
                        help="Path to govcloud_config.yaml")
    parser.add_argument("--setup", action="store_true",
                        help="Verify and apply patches only")
    parser.add_argument("--generate", action="store_true",
                        help="Generate test cases from Excel")
    parser.add_argument("--evaluate", action="store_true",
                        help="Run ADK evaluation")
    parser.add_argument("--judge", action="store_true",
                        help="Run LLM-as-Judge evaluation")
    parser.add_argument("--report", action="store_true",
                        help="Show results report")
    parser.add_argument("--all", action="store_true",
                        help="Generate + Evaluate + Judge + Report")
    parser.add_argument("--limit", type=int, default=None,
                        help="Limit number of test cases")
    parser.add_argument("--all-questions", action="store_true",
                        help="Include all questions (not just MVP)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Show what patches would be applied without making changes")
    parser.add_argument("--verbose", "-v", action="store_true",
                        help="Show detailed output")

    args = parser.parse_args()

    # Load config
    config = load_yaml_config(Path(args.config))

    if not any([args.setup, args.generate, args.evaluate, args.judge, args.report, args.all]):
        parser.print_help()
        sys.exit(0)

    exit_code = 0

    if args.setup:
        exit_code = cmd_setup(config, args)
        if exit_code != 0:
            sys.exit(exit_code)

    if args.all or args.generate:
        exit_code = cmd_generate(config, args)
        if exit_code != 0:
            sys.exit(exit_code)

    if args.all or args.evaluate:
        exit_code = cmd_evaluate(config, args)
        if exit_code != 0:
            sys.exit(exit_code)

    if args.all or args.judge:
        exit_code = cmd_judge(config, args)
        if exit_code != 0:
            sys.exit(exit_code)

    if args.all or args.report:
        exit_code = cmd_report(config, args)

    sys.exit(exit_code)


if __name__ == "__main__":
    main()
